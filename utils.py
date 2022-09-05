from __future__ import print_function

import functools
import os

from datetime import datetime
from openpyxl import load_workbook, Workbook
from PyPDF2 import PdfFileReader, PdfFileWriter


# Max size of pdf file - bytes
FILE_MAX_SIZE = 10485760

DOWNLOAD_DIR = 'downloads'

TF = '%Y-%m-%d %H-%M-%S'


def get_file_size(file):
    pos = file.tell()
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(pos)
    return size


def validate_and_save_pdf(pdf_path: str):
    if not os.path.exists(pdf_path):
        raise ValueError('File does not exists')
    if not pdf_path.endswith('.pdf'):
        raise ValueError('Wrong file format')

    with open(pdf_path, 'rb') as file:
        file_size = get_file_size(file)
        if file_size > FILE_MAX_SIZE:
            raise ValueError('Maximum file size: 10 Mb')
        try:
            pdf = PdfFileReader(file)
        except Exception:
            raise ValueError('You must upload a valid PDF file')
        if pdf.is_encrypted:
            raise ValueError('Your pdf file is encrypted')

        pdf_writer = PdfFileWriter()
        for page in range(pdf.getNumPages()):
            pdf_writer.addPage(pdf.getPage(page))

        file_name = f'{DOWNLOAD_DIR}/{os.path.basename(pdf_path)}'
        if os.path.exists(file_name):
            file_name = f'{file_name}_{datetime.now()}'

        with open(file_name, 'wb') as output_pdf:
            pdf_writer.write(output_pdf)
            output_pdf.close()
        file.close()
    return f'File has been successfully checked and saved ' \
           f'to folder "{DOWNLOAD_DIR}"'


def get_xlsx_data(file_path: str):
    """
    Returns dict for processing pdf files in main module.
    key: "folder_name" reads from first_column of xlsx file;
    values: "urls" reads from second column
    """
    wb = load_workbook(file_path)
    sheet = wb.active
    folders = [folder_name.value for folder_name in sheet['A']]
    urls = [url.value.strip().split(';') for url in sheet['B']]
    xlsx_dict = {folder_name: urls[index]
                 for index, folder_name in enumerate(folders)}
    wb.close()
    return xlsx_dict


def make_xlsx_report(func):
    """
    create report in xlsx format, depends on structure of dict
    which returns wrapped func.
    """
    @functools.wraps(func)
    def wrapper_make_xlsx_report(*args, **kwargs):
        value = func(*args, **kwargs)
        if value:
            wb = Workbook()
            ws = wb.create_sheet(
                f'report_{datetime.now().strftime(TF)}')
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 50
            ws.cell(row=1, column=1, value='Folder name')
            ws.cell(row=1, column=2, value='File name')
            ws.cell(row=1, column=3, value='File id')
            ws.cell(row=1, column=4, value='Status')
            counter = 2
            for folder_name in value.keys():
                values = value[folder_name]
                for row in range(len(values)):
                    ws.cell(row=counter, column=1, value=folder_name)
                    ws.cell(row=counter, column=2, value=values[row][0])
                    ws.cell(row=counter, column=3, value=values[row][1])
                    ws.cell(row=counter, column=4, value=values[row][2])
                    counter += 1
            wb.save('report.xlsx')
        return value
    return wrapper_make_xlsx_report
